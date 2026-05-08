from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.contrib import messages
from django.db.models import Sum
from django.utils import timezone
from .models import Applicant, Application, Specialization, AdmissionCampaign, Attachment
import csv
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import matplotlib # type: ignore
matplotlib.use('Agg')

import matplotlib.pyplot as plt # type: ignore
import io
import base64
from django.db.models import Count, Avg
import io
from django.http import FileResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, Table, TableStyle, SimpleDocTemplate
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
# 1. Для проверки прав доступа (ошибка user_passes_test)
from django.contrib.auth.decorators import user_passes_test

# 2. Для работы с PDF (ошибка TA_RIGHT)
from reportlab.lib.enums import TA_RIGHT

# 3. Для модели приказов (ошибка EnrollmentOrder)
# Проверьте, что модель EnrollmentOrder реально создана в вашем models.py
from .models import Applicant, Application, Specialization, AdmissionCampaign, Attachment, EnrollmentOrder, ExamResult



def home(request):
    """Главная страница с информацией о специальностях"""
    specializations = Specialization.objects.filter(is_active=True)
    active_campaign = AdmissionCampaign.objects.filter(is_active=True).first()
    context = {
        'specializations': specializations,
        'active_campaign': active_campaign,
    }
    return render(request, 'applicants/home.html', context)


def register_view(request):
    """Регистрация нового абитуриента"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        password2 = request.POST.get('password2')
        email = request.POST.get('email')
        last_name = request.POST.get('last_name')
        first_name = request.POST.get('first_name')
        patronymic = request.POST.get('patronymic', '')
        birth_date = request.POST.get('birth_date')
        sex = request.POST.get('sex')
        phone = request.POST.get('phone')

        # Валидация
        if password != password2:
            messages.error(request, 'Пароли не совпадают')
            return render(request, 'applicants/register.html')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Пользователь с таким логином уже существует')
            return render(request, 'applicants/register.html')

        # Создание пользователя
        user = User.objects.create_user(
            username=username,
            password=password,
            email=email,
            last_name=last_name,
            first_name=first_name
        )

        # Создание профиля абитуриента
        applicant = Applicant.objects.create(
            user=user,
            patronymic=patronymic,
            birth_date=birth_date,
            sex=sex,
            phone=phone
        )

        # Добавление в группу "Абитуриенты"
        applicant_group, _ = Group.objects.get_or_create(name='Applicants')
        user.groups.add(applicant_group)

        login(request, user)
        messages.success(request, 'Регистрация прошла успешно!')
        return redirect('dashboard')

    return render(request, 'applicants/register.html')


def login_view(request):
    """Вход в систему"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            # Проверяем, есть ли профиль абитуриента
            if hasattr(user, 'applicant_profile'):
                return redirect('dashboard')
            else:
                return redirect('admin:index')
        else:
            messages.error(request, 'Неверный логин или пароль')

    return render(request, 'applicants/login.html')


def logout_view(request):
    """Выход из системы"""
    logout(request)
    return redirect('home')


@login_required
def dashboard(request):
    """Личный кабинет абитуриента"""
    try:
        applicant = request.user.applicant_profile
    except Applicant.DoesNotExist:
        messages.error(request, 'Профиль абитуриента не найден')
        return redirect('home')

    applications = Application.objects.filter(applicant=applicant).annotate(
        total_score=Sum('exam_results__score')
    )
    attachments = Attachment.objects.filter(applicant=applicant)

    context = {
        'applicant': applicant,
        'applications': applications,
        'attachments': attachments,
    }
    return render(request, 'applicants/dashboard.html', context)


@login_required
def create_application(request):
    """Подача заявления с баллами ЕГЭ"""
    try:
        applicant = request.user.applicant_profile
    except Applicant.DoesNotExist:
        messages.error(request, 'Профиль абитуриента не найден')
        return redirect('home')

    active_campaign = AdmissionCampaign.objects.filter(is_active=True).first()
    if not active_campaign:
        messages.error(request, 'Нет активной приемной кампании')
        return redirect('dashboard')

    if request.method == 'POST':
        specialization_id = request.POST.get('specialization')
        priority = request.POST.get('priority')
        
        # Баллы ЕГЭ
        score_russian = request.POST.get('score_russian')
        score_math = request.POST.get('score_math')
        extra_subject = request.POST.get('extra_subject')
        score_extra = request.POST.get('score_extra')

        specialization = get_object_or_404(Specialization, id=specialization_id, is_active=True)

        # Проверка уникальности
        if Application.objects.filter(
            applicant=applicant,
            campaign=active_campaign,
            priority=priority
        ).exists():
            messages.error(request, f'У вас уже есть заявление с приоритетом {priority}')
            return redirect('create_application')

        if Application.objects.filter(
            applicant=applicant,
            campaign=active_campaign,
            specialization=specialization
        ).exists():
            messages.error(request, 'Вы уже подали заявление на эту специальность')
            return redirect('create_application')

        # Создаём заявление
        application = Application.objects.create(
            applicant=applicant,
            campaign=active_campaign,
            specialization=specialization,
            priority=int(priority),
            status='submitted'
        )

        # Сохраняем баллы ЕГЭ
        if score_russian:
            ExamResult.objects.create(
                application=application,
                subject='Русский язык',
                score=int(score_russian),
                is_ege=True
            )
        
        if score_math:
            ExamResult.objects.create(
                application=application,
                subject='Математика',
                score=int(score_math),
                is_ege=True
            )
        
        if extra_subject and score_extra:
            ExamResult.objects.create(
                application=application,
                subject=extra_subject,
                score=int(score_extra),
                is_ege=True
            )

        messages.success(request, 'Заявление успешно подано! Баллы сохранены.')
        return redirect('dashboard')

    specializations = Specialization.objects.filter(is_active=True)
    existing_priorities = Application.objects.filter(
        applicant=applicant,
        campaign=active_campaign
    ).values_list('priority', flat=True)

    context = {
        'specializations': specializations,
        'existing_priorities': list(existing_priorities),
    }
    return render(request, 'applicants/create_application.html', context)


@login_required
def application_detail(request, pk):
    """Детали заявления"""
    try:
        applicant = request.user.applicant_profile
    except Applicant.DoesNotExist:
        return redirect('home')

    application = get_object_or_404(Application, pk=pk, applicant=applicant)
    exam_results = application.exam_results.all()
    total_score = exam_results.aggregate(s=Sum('score'))['s'] or 0

    context = {
        'application': application,
        'exam_results': exam_results,
        'total_score': total_score,
    }
    return render(request, 'applicants/application_detail.html', context)


@login_required
def sign_agreement(request, pk):
    """Подписание согласия на зачисление"""
    try:
        applicant = request.user.applicant_profile
    except Applicant.DoesNotExist:
        return redirect('home')

    application = get_object_or_404(Application, pk=pk, applicant=applicant)

    if request.method == 'POST':
        application.is_agreement_signed = True
        application.status = 'agreement_signed'
        application.agreement_date = timezone.now()
        application.save()
        messages.success(request, 'Согласие на зачисление подписано!')
        return redirect('application_detail', pk=pk)

    context = {'application': application}
    return render(request, 'applicants/sign_agreement.html', context)

def rating_view(request, specialization_id):
    """Просмотр рейтинга по специальности"""
    specialization = get_object_or_404(Specialization, id=specialization_id, is_active=True)
    
    # Получаем минимальный проходной балл
    min_score = specialization.min_score if hasattr(specialization, 'min_score') else 100
    
    # Отладка в консоли
    print(f"\n=== РЕЙТИНГ: {specialization.name} ===")
    print(f"Мин. балл: {min_score}")
    print(f"Бюджетных мест: {specialization.budget_places}")
    print(f"Платных мест: {specialization.paid_places}")

    # Получаем ВСЕ заявления с подписанным согласием
    applications = Application.objects.filter(
        specialization=specialization,
        is_agreement_signed=True
    ).annotate(
        total_score=Sum('exam_results__score')
    ).order_by('-total_score', 'submission_date')
    
    print(f"Всего заявлений с согласием: {applications.count()}")
    for app in applications:
        total = app.total_score or 0
        status = "ПРОХОДИТ" if total >= min_score else "НЕ ПРОХОДИТ"
        print(f"  - {app.applicant.full_name}: {total} баллов → {status} (мин. {min_score})")

    context = {
        'specialization': specialization,
        'applications': applications,
        'min_score': min_score,
    }
    return render(request, 'applicants/rating.html', context)


@login_required
def profile_view(request):
    """Редактирование профиля"""
    try:
        applicant = request.user.applicant_profile
    except Applicant.DoesNotExist:
        return redirect('home')

    if request.method == 'POST':
        applicant.phone = request.POST.get('phone', applicant.phone)
        applicant.passport_series = request.POST.get('passport_series', applicant.passport_series)
        applicant.passport_number = request.POST.get('passport_number', applicant.passport_number)
        applicant.address = request.POST.get('address', applicant.address)
        applicant.snils = request.POST.get('snils', applicant.snils)
        applicant.save()

        request.user.last_name = request.POST.get('last_name', request.user.last_name)
        request.user.first_name = request.POST.get('first_name', request.user.first_name)
        request.user.email = request.POST.get('email', request.user.email)
        request.user.save()

        applicant.patronymic = request.POST.get('patronymic', applicant.patronymic)
        applicant.save()

        messages.success(request, 'Профиль обновлен')
        return redirect('profile')

    context = {'applicant': applicant}
    return render(request, 'applicants/profile.html', context)


@login_required
def upload_document(request):
    """Загрузка документов"""
    try:
        applicant = request.user.applicant_profile
    except Applicant.DoesNotExist:
        return redirect('home')

    if request.method == 'POST' and request.FILES.get('file'):
        doc_type = request.POST.get('doc_type', 'other')
        Attachment.objects.create(
            applicant=applicant,
            file=request.FILES['file'],
            doc_type=doc_type
        )
        messages.success(request, 'Документ загружен')
        return redirect('dashboard')

    return redirect('dashboard')

def export_rating_excel(request, specialization_id):
    """Экспорт рейтинга в Excel"""
    specialization = get_object_or_404(Specialization, id=specialization_id)
    
    applications = Application.objects.filter(
        specialization=specialization,
        is_agreement_signed=True
    ).annotate(
        total_score=Sum('exam_results__score')
    ).order_by('-total_score', 'submission_date')
    
    # Создаем рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Рейтинг {specialization.code}"
    
    # Заголовок
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Рейтинг абитуриентов: {specialization.name} ({specialization.code})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Информация о местах
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Бюджетных мест: {specialization.budget_places} | Платных мест: {specialization.paid_places}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Шапка таблицы
    headers = ['Место', 'ФИО абитуриента', 'Сумма баллов', 'СНИЛС', 'Дата подачи']
    header_style = Font(bold=True, color='FFFFFF')
    header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Данные
    green_fill = openpyxl.styles.PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    for index, app in enumerate(applications, 1):
        row = index + 4
        ws.cell(row=row, column=1, value=index)
        ws.cell(row=row, column=2, value=app.applicant.full_name)
        ws.cell(row=row, column=3, value=app.total_score or 0)
        ws.cell(row=row, column=4, value=app.applicant.snils)
        ws.cell(row=row, column=5, value=app.submission_date.strftime('%d.%m.%Y'))
        
        # Зеленый для бюджетных мест
        if index <= specialization.budget_places:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = green_fill
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Сохраняем в HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=rating_{specialization.code}.xlsx'
    wb.save(response)
    
    return response

def statistics_view(request):
    """Страница статистики (для председателя ПК)"""
    active_campaign = AdmissionCampaign.objects.filter(is_active=True).first()
    
    if not active_campaign:
        messages.error(request, 'Нет активной приемной кампании')
        return redirect('home')
    
    # Статистика по специальностям
    specializations_stats = []
    for spec in Specialization.objects.filter(is_active=True):
        total_apps = Application.objects.filter(
            campaign=active_campaign, 
            specialization=spec
        ).count()
        with_agreement = Application.objects.filter(
            campaign=active_campaign,
            specialization=spec,
            is_agreement_signed=True
        ).count()
        avg_score = Application.objects.filter(
            campaign=active_campaign,
            specialization=spec,
            is_agreement_signed=True
        ).annotate(
            total=Sum('exam_results__score')
        ).aggregate(avg=Avg('total'))['avg']
        
        specializations_stats.append({
            'name': spec.name,
            'code': spec.code,
            'total_applications': total_apps,
            'with_agreement': with_agreement,
            'budget_places': spec.budget_places,
            'competition': round(total_apps / spec.budget_places, 2) if spec.budget_places > 0 else 0,
            'avg_score': round(avg_score, 1) if avg_score else 0,
        })
    
    # Генерация графика "Распределение по специальностям"
    fig, ax = plt.subplots(figsize=(10, 6))
    names = [s['code'] for s in specializations_stats]
    total_apps = [s['total_applications'] for s in specializations_stats]
    with_agr = [s['with_agreement'] for s in specializations_stats]
    
    x = range(len(names))
    width = 0.35
    ax.bar([i - width/2 for i in x], total_apps, width, label='Всего заявлений', color='#4472C4')
    ax.bar([i + width/2 for i in x], with_agr, width, label='С согласием', color='#70AD47')
    ax.set_xticks(x)
    ax.set_xticklabels(names, fontsize=8)
    ax.legend()
    ax.set_title('Статистика подачи заявлений по специальностям', fontsize=14, pad=20)
    ax.set_ylabel('Количество', fontsize=12)
    ax.grid(axis='y', alpha=0.3)
    
    # Сохранение графика в base64 для отображения
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', bbox_inches='tight', dpi=100)
    buffer.seek(0)
    chart_base64 = base64.b64encode(buffer.getvalue()).decode()
    plt.close()
    
    # Общая статистика
    total_applicants = Applicant.objects.filter(
        applications__campaign=active_campaign
    ).distinct().count()
    
    total_applications = Application.objects.filter(campaign=active_campaign).count()
    
    context = {
        'specializations_stats': specializations_stats,
        'chart': chart_base64,
        'total_applicants': total_applicants,
        'total_applications': total_applications,
        'campaign': active_campaign,
    }
    
    return render(request, 'applicants/statistics.html', context)

# ========== ПРИКАЗЫ О ЗАЧИСЛЕНИИ ==========

def get_available_priorities(applicant, campaign):
    """Вспомогательная функция: получение доступных приоритетов"""
    existing = Application.objects.filter(
        applicant=applicant,
        campaign=campaign
    ).values_list('priority', flat=True)
    return [p for p in [1, 2, 3] if p not in existing]


def enrollment_orders_list(request):
    """Страница со списком приказов"""
    active_campaign = AdmissionCampaign.objects.filter(is_active=True).first()
    
    if not active_campaign:
        messages.error(request, 'Нет активной приемной кампании')
        return redirect('home')
    
    orders = EnrollmentOrder.objects.filter(campaign=active_campaign).order_by('-order_date')
    specializations = Specialization.objects.filter(is_active=True)
    
    context = {
        'orders': orders,
        'specializations': specializations,
        'campaign': active_campaign,
    }
    return render(request, 'applicants/enrollment_orders.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def create_order(request):
    """Создание приказа о зачислении"""
    if request.method != 'POST':
        return redirect('enrollment_orders')
    
    active_campaign = AdmissionCampaign.objects.filter(is_active=True).first()
    if not active_campaign:
        messages.error(request, 'Нет активной приемной кампании')
        return redirect('enrollment_orders')
    
    order_number = request.POST.get('order_number')
    order_date = request.POST.get('order_date')
    specialization_id = request.POST.get('specialization')
    
    if not all([order_number, order_date, specialization_id]):
        messages.error(request, 'Заполните все поля')
        return redirect('enrollment_orders')
    
    specialization = get_object_or_404(Specialization, id=specialization_id)
    
    # Минимальный балл
    min_score = getattr(specialization, 'min_score', 0)
    
    print(f"\n{'='*50}")
    print(f"СОЗДАНИЕ ПРИКАЗА")
    print(f"Специальность: {specialization.name}")
    print(f"Мин. балл: {min_score}")
    print(f"Бюджетных мест: {specialization.budget_places}")
    print(f"Платных мест: {specialization.paid_places}")
    
    # Все заявления с согласием
    all_apps = Application.objects.filter(
        specialization=specialization,
        campaign=active_campaign,
        is_agreement_signed=True,
    ).annotate(
        total_score=Sum('exam_results__score')
    ).order_by('-total_score', 'submission_date')
    
    # Отбираем ТОЛЬКО тех, кто прошёл минимальный порог
    qualified_apps = [app for app in all_apps if (app.total_score or 0) >= min_score]
    
    print(f"Всего с согласием: {all_apps.count()}")
    print(f"Прошли порог ({min_score}): {len(qualified_apps)}")
    
    # Распределяем: бюджет, потом платное
    budget_list = []
    paid_list = []
    
    for app in qualified_apps:
        if len(budget_list) < specialization.budget_places:
            budget_list.append(app)
        elif len(paid_list) < specialization.paid_places:
            paid_list.append(app)
    
    enrolled = budget_list + paid_list
    
    print(f"Бюджет: {len(budget_list)} чел.")
    for app in budget_list:
        print(f"  - {app.applicant.full_name}: {app.total_score or 0} баллов")
    print(f"Платное: {len(paid_list)} чел.")
    for app in paid_list:
        print(f"  - {app.applicant.full_name}: {app.total_score or 0} баллов")
    
    if not enrolled:
        messages.error(request, f'Нет абитуриентов с баллами >= {min_score}')
        return redirect('enrollment_orders')
    
    # ID для обновления
    budget_ids = [app.id for app in budget_list]
    paid_ids = [app.id for app in paid_list]
    all_enrolled_ids = budget_ids + paid_ids
    all_applicant_ids = list(set(app.applicant_id for app in enrolled))
    
    # Создаём приказ
    order = EnrollmentOrder.objects.create(
        campaign=active_campaign,
        order_number=order_number,
        order_date=order_date,
        created_by=request.user
    )
    order.applications.set(enrolled)
    
    # Обновляем статусы
    Application.objects.filter(id__in=all_enrolled_ids).update(status='accepted')
    Applicant.objects.filter(id__in=all_applicant_ids).update(status='accepted')
    
    # Сохраняем в сессии для PDF
    request.session['budget_ids'] = budget_ids
    request.session['paid_ids'] = paid_ids
    
    messages.success(
        request,
        f'Приказ №{order_number} создан! Бюджет: {len(budget_list)} чел., Платное: {len(paid_list)} чел.'
    )
    print(f"{'='*50}\n")
    return redirect('enrollment_orders')


def enrollment_orders_list(request):
    """Страница со списком приказов"""
    active_campaign = AdmissionCampaign.objects.filter(is_active=True).first()
    
    if not active_campaign:
        messages.error(request, 'Нет активной приемной кампании')
        return redirect('home')
    
    # Если сотрудник — видит все приказы
    if request.user.is_staff:
        orders = EnrollmentOrder.objects.filter(campaign=active_campaign).order_by('-order_date')
    else:
        # Абитуриент видит только приказы, где он есть
        try:
            applicant = request.user.applicant_profile
            orders = EnrollmentOrder.objects.filter(
                campaign=active_campaign,
                applications__applicant=applicant
            ).distinct().order_by('-order_date')
        except Applicant.DoesNotExist:
            orders = EnrollmentOrder.objects.none()
    
    specializations = Specialization.objects.filter(is_active=True)
    
    context = {
        'orders': orders,
        'specializations': specializations,
        'campaign': active_campaign,
    }
    return render(request, 'applicants/enrollment_orders.html', context)

@login_required
def download_order_pdf(request, order_id):
    """Генерация PDF-приказа (для сотрудников — полный, для абитуриента — выписка)"""
    order = get_object_or_404(EnrollmentOrder, id=order_id)
    
    # Проверка доступа
    if not request.user.is_staff:
        try:
            applicant = request.user.applicant_profile
            if not order.applications.filter(applicant=applicant).exists():
                messages.error(request, 'У вас нет доступа к этому приказу')
                return redirect('enrollment_orders')
        except Applicant.DoesNotExist:
            messages.error(request, 'Доступ запрещён')
            return redirect('home')
    
    # Получаем заявления
    if request.user.is_staff:
        applications = order.applications.all().annotate(
            total_score=Sum('exam_results__score')
        ).order_by('-total_score')
    else:
        applicant = request.user.applicant_profile
        applications = order.applications.filter(applicant=applicant).annotate(
            total_score=Sum('exam_results__score')
        ).order_by('-total_score')
    
    print(f"\nГенерация PDF для приказа №{order.order_number}")
    print(f"Заявлений в PDF: {applications.count()}")
    
    # ===== ПОДКЛЮЧАЕМ РУССКИЙ ШРИФТ (Arial) =====
    try:
        # Windows
        pdfmetrics.registerFont(TTFont('ArialRegular', 'C:/Windows/Fonts/arial.ttf'))
        pdfmetrics.registerFont(TTFont('ArialBold', 'C:/Windows/Fonts/arialbd.ttf'))
        print("Шрифты Arial загружены (Windows)")
    except:
        try:
            # Mac
            pdfmetrics.registerFont(TTFont('ArialRegular', '/Library/Fonts/Arial.ttf'))
            pdfmetrics.registerFont(TTFont('ArialBold', '/Library/Fonts/Arial Bold.ttf'))
            print("Шрифты Arial загружены (Mac)")
        except:
            try:
                # Linux
                pdfmetrics.registerFont(TTFont('ArialRegular', '/usr/share/fonts/truetype/msttcorefonts/Arial.ttf'))
                pdfmetrics.registerFont(TTFont('ArialBold', '/usr/share/fonts/truetype/msttcorefonts/Arial_Bold.ttf'))
                print("Шрифты Arial загружены (Linux)")
            except:
                # Если вообще ничего не нашлось - используем стандартный
                print("ВНИМАНИЕ: Arial не найден, используем Courier")
                pdfmetrics.registerFont(TTFont('ArialRegular', 'Courier'))
                pdfmetrics.registerFont(TTFont('ArialBold', 'Courier-Bold'))
    
    from reportlab.pdfbase.pdfmetrics import registerFontFamily
    registerFontFamily('ArialFamily', normal='ArialRegular', bold='ArialBold')
    print("Семейство шрифтов зарегистрировано")
    # ====================================
    
    # Создаем PDF
    buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=25*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    # Стили - используем ТОЧНО ТЕ ЖЕ ИМЕНА что при регистрации
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        fontName='ArialBold',        # ← ТОЧНО как при регистрации
        fontSize=14,
        alignment=TA_CENTER,
        spaceAfter=10*mm,
        textColor=colors.HexColor('#1a237e'),
    )
    
    subtitle_style = ParagraphStyle(
        'SubTitle',
        fontName='ArialRegular',     # ← ТОЧНО как при регистрации
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=5*mm,
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        fontName='ArialRegular',     # ← ТОЧНО как при регистрации
        fontSize=10,
        leading=14,
        spaceAfter=3*mm,
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        fontName='ArialBold',        # ← ТОЧНО как при регистрации
        fontSize=10,
        textColor=colors.white,
        alignment=TA_CENTER,
    )
    
    table_cell_style = ParagraphStyle(
        'TableCell',
        fontName='ArialRegular',     # ← ТОЧНО как при регистрации
        fontSize=9,
        leading=12,
        alignment=TA_CENTER,
    )
    
    # Формируем содержимое
    elements = []
    
    # Шапка
    elements.append(Paragraph(
        'МИНИСТЕРСТВО ОБРАЗОВАНИЯ РОССИЙСКОЙ ФЕДЕРАЦИИ',
        subtitle_style
    ))
    elements.append(Paragraph(
        'Колледж мировой экономики и передовых технологий',
        subtitle_style
    ))
    
    if request.user.is_staff:
        # Полный текст для сотрудников
        spec_name = applications.first().specialization.name if applications.exists() else ''
        
        # Получаем ID бюджетников и платников из сессии
        budget_ids = request.session.get('budget_ids', [])
        paid_ids = request.session.get('paid_ids', [])
        
        elements.append(Paragraph(
            f'В соответствии с Правилами приема и на основании решения Приемной комиссии '
            f'(протокол от {order.order_date.strftime("%d.%m.%Y")} г.) <b>ПРИКАЗЫВАЮ:</b>',
            normal_style
        ))
        elements.append(Paragraph(
            f'1. Зачислить с {order.order_date.strftime("%d.%m.%Y")} г. в число студентов первого курса '
            f'по специальности <b>{spec_name}</b> следующих абитуриентов, '
            f'успешно прошедших вступительные испытания:',
            normal_style
        ))
        
        # Таблица со всеми абитуриентами
        table_data = [
            [
                Paragraph('№ п/п', table_header_style),
                Paragraph('Фамилия, Имя, Отчество', table_header_style),
                Paragraph('Сумма баллов', table_header_style),
                Paragraph('Вид финансирования', table_header_style),
            ]
        ]
        
        for index, app in enumerate(applications, 1):
            # Определяем вид финансирования
            if app.id in budget_ids:
                finance_type = 'Бюджет'
            elif app.id in paid_ids:
                finance_type = 'Платное'
            else:
                finance_type = 'Бюджет'  # По умолчанию
            
            table_data.append([
                Paragraph(str(index), table_cell_style),
                Paragraph(app.applicant.full_name, table_cell_style),
                Paragraph(str(app.total_score or 0), table_cell_style),
                Paragraph(finance_type, table_cell_style),
            ])
        
        col_widths = [15*mm, 90*mm, 30*mm, 30*mm]
        
    else:
        # Текст выписки для абитуриента
        app = applications.first()
        
        # Определяем вид финансирования
        budget_ids = request.session.get('budget_ids', [])
        paid_ids = request.session.get('paid_ids', [])
        
        if app.id in budget_ids:
            finance_type = 'бюджетную'
        elif app.id in paid_ids:
            finance_type = 'платную'
        else:
            finance_type = 'бюджетную'
        
        elements.append(Paragraph(
            f'На основании приказа №{order.order_number} от {order.order_date.strftime("%d.%m.%Y")} г. '
            f'по Колледжу мировой экономики и передовых технологий',
            normal_style
        ))
        elements.append(Paragraph(
            f'<b>{app.applicant.full_name}</b> зачислен(а) с {order.order_date.strftime("%d.%m.%Y")} г. '
            f'в число студентов 1 курса по специальности <b>{app.specialization.name}</b> '
            f'({app.specialization.code}) на <b>{finance_type} основу</b>.',
            normal_style
        ))
        elements.append(Paragraph(
            f'Сумма баллов: <b>{app.total_score or 0}</b>',
            normal_style
        ))
        elements.append(Paragraph('<br/>', normal_style))
        elements.append(Paragraph(
            'Выписка дана для предоставления по месту требования.',
            normal_style
        ))
        
        table_data = [
            [
                Paragraph('Фамилия, Имя, Отчество', table_header_style),
                Paragraph('Специальность', table_header_style),
                Paragraph('Сумма баллов', table_header_style),
                Paragraph('Основа', table_header_style),
            ]
        ]
        table_data.append([
            Paragraph(app.applicant.full_name, table_cell_style),
            Paragraph(f'{app.specialization.code} — {app.specialization.name}', table_cell_style),
            Paragraph(str(app.total_score or 0), table_cell_style),
            Paragraph('Бюджет', table_cell_style),
        ])
        
        col_widths = [70*mm, 60*mm, 25*mm, 20*mm]
    
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fa')]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(table)
    elements.append(Paragraph('<br/><br/>', normal_style))
    
    if request.user.is_staff:
        elements.append(Paragraph('Директор колледжа: ____________________', normal_style))
        elements.append(Paragraph('Секретарь приемной комиссии: ____________________', normal_style))
    else:
        elements.append(Paragraph('Секретарь приемной комиссии: ____________________', normal_style))
        elements.append(Paragraph('<br/>', normal_style))
        elements.append(Paragraph('М.П.', normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    print(f"PDF сгенерирован успешно!\n")
    
    filename = f'order_{order.order_number}.pdf' if request.user.is_staff else f'vypiska_{order.order_number}.pdf'
    
    return FileResponse(
        buffer,
        as_attachment=True,
        filename=filename,
        content_type='application/pdf'
    )
@login_required
def upload_document_page(request):
    """Страница загрузки документов"""
    try:
        applicant = request.user.applicant_profile
    except Applicant.DoesNotExist:
        return redirect('home')
    
    if request.method == 'POST' and request.FILES.get('file'):
        doc_type = request.POST.get('doc_type', 'other')
        Attachment.objects.create(
            applicant=applicant,
            file=request.FILES['file'],
            doc_type=doc_type
        )
        messages.success(request, 'Документ успешно загружен!')
        return redirect('dashboard')
    
    attachments = Attachment.objects.filter(applicant=applicant)
    
    context = {
        'attachments': attachments,
    }
    return render(request, 'applicants/upload_document.html', context)